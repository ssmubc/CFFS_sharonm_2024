import pandas as pd
import openpyxl
import pytest
from datetime import datetime
import xlrd
from pyxlsb import open_workbook
from generate_menu_list import OK_list

def identify_missing_items(excel_file, adict):
    absent_list = []
    recipe_list = list(adict.keys())

    def get_sheet_names(excel_file):
        xlsx = openpyxl.load_workbook(f"/Users/jennylee/CFFS-PyCharm/data/menu_list/{excel_file}.xlsx")
        sheetnames = xlsx.sheetnames
        sheetnames = sheetnames[0:len(sheetnames) - 3]
        return sheetnames

    for sheet in get_sheet_names(excel_file):
        menu_file = pd.read_excel(f"/Users/jennylee/CFFS-PyCharm/data/menu_list/{excel_file}.xlsx", sheet_name=sheet)
        menu_file = menu_file.iloc[:, 1:17]

        for item in list(menu_file):
            item_list = menu_file[item].unique().tolist()
            for item_2 in item_list:
                if (item_2 not in recipe_list) & (item_2 not in absent_list):
                    absent_list.append(item_2)

    absent_list = [item for item in absent_list if not (pd.isnull(item) == True)]
    return absent_list

if __name__ == '__main__':
    print(identify_missing_items("OK", OK_list))
    print(len(identify_missing_items("OK", OK_list)))

    with open("menu_list/OK_list.txt", "w") as output:
        output.write(str(identify_missing_items("OK", OK_list)))